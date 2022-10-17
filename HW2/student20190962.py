import openpyxl
import operator

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

scores = dict()
row_num = 1
for row in ws:
	if row_num != 1:
		total = ws.cell(row = row_num, column = 3).value * 0.3
		total += ws.cell(row = row_num, column = 4).value * 0.35
		total += ws.cell(row = row_num, column = 5).value * 0.34
		total += ws.cell(row = row_num, column = 6).value * 0.01
		ws.cell(row = row_num, column = 7).value = total
		scores[row_num] = total
	row_num += 1

rank = sorted(scores.items(), key=operator.itemgetter(1), reverse=True)
rankList = []

for ranks in rank:
	rankList.append(list(ranks))

count = int(1)
scoreList = []

for i in range(len(rankList)):
	if (i == 0):
		count = 1
	elif(rankList[i][1] != rankList[i-1][1]):
		count = count + 1
	scoreList.append(count)

finalscore=[1]
i=1

while i < len(scoreList):
	if scoreList.count(scoreList[i]) > 1:
		k = i
		for j in range(scoreList.count(scoreList[k])):
			finalscore.append(finalscore[k-1]+scoreList.count(scoreList[k]))
			i += 1
	else:
		finalscore.append(finalscore[i-1]+1)
		i += 1

A = int(len(finalscore)*0.3)
AP = int(A/2)
B = int(len(finalscore)*0.7)
BP = int((B-A)/2) + A
CP = B + int((len(finalscore)-B)/2)

for s in range(len(finalscore)):
	fs = finalscore[s]
	if fs <= AP:
		grade = 'A+'
	elif fs <= A:
		grade = 'A0'
	elif fs <= BP:
		grade = 'B+'
	elif fs <= B:
		grade = 'B0'
	elif fs <= CP:
		grade = 'C+'
	else:
		grade = 'C0'
	ws.cell(row=rankList[s][0], column=8).value = grade

wb.save("student.xlsx")


